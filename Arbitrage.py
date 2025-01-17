import requests
import xlsxwriter
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, numbers

API_KEY = "Your API KEY"


SPORT = 'upcoming' # use the sport_key from the /sports endpoint below, or use 'upcoming' to see the next 8 games across all sports

REGIONS = 'us' # uk | us | eu | au. Multiple can be specified if comma delimited

MARKETS = 'h2h' # h2h | spreads | totals. Multiple can be specified if comma delimited

ODDS_FORMAT = 'decimal' # decimal | american

DATE_FORMAT = 'iso' # iso | unix

BET_SIZE = 100

odds_response = requests.get(
    f'https://api.the-odds-api.com/v4/sports/{SPORT}/odds',
    params={
        'api_key': API_KEY,
        'regions': REGIONS,
        'markets': MARKETS,
        'oddsFormat': ODDS_FORMAT,
        'dateFormat': DATE_FORMAT,
    }
).json()

# odds_response
#print(odds_response)

BOOKMAKER_INDEX = 0
NAME_INDEX = 1
ODDS_INDEX = 2
FIRST = 0

class Event:
    def __init__(self, data):
        self.data = data
        self.sport_key = data['sport_key']
        self.id = data['id']
        
    def find_best_odds(self):
        # number of possible outcomes for a sporting event
        num_outcomes = len(self.data['bookmakers'][FIRST]['markets'][FIRST]['outcomes'])
        self.num_outcomes = num_outcomes
        

        # finding the best odds for each outcome in each event
        best_odds = [[None, None, float('-inf')] for _ in range(num_outcomes)]
        # [Bookmaker, Name, Price]

        bookmakers = event.data['bookmakers']
        for index, bookmaker in enumerate(bookmakers):

            # determing the odds offered by each bookmaker
            for outcome in range(num_outcomes):

                # determining if any of the bookmaker odds are better than the current best odds
                bookmaker_odds = float(bookmaker['markets'][FIRST]['outcomes'][outcome]['price'])
                current_best_odds = best_odds[outcome][ODDS_INDEX]

                if bookmaker_odds > current_best_odds:
                    best_odds[outcome][BOOKMAKER_INDEX] = bookmaker['title']
                    best_odds[outcome][NAME_INDEX] = bookmaker['markets'][FIRST]['outcomes'][outcome]['name']
                    best_odds[outcome][ODDS_INDEX] = bookmaker_odds
                    
        self.best_odds = best_odds
        return best_odds
    
    def arbitrage(self):
        total_arbitrage_percentage = 0
        for odds in self.best_odds:
            total_arbitrage_percentage += (1.0 / odds[ODDS_INDEX])
            
        self.total_arbitrage_percentage = total_arbitrage_percentage
        self.expected_earnings = (BET_SIZE / total_arbitrage_percentage) - BET_SIZE
        
        # if the sum of the reciprocals of the odds is less than 1, there is opportunity for arbitrage
        if total_arbitrage_percentage < 1:
            return True
        return False
    
    # converts decimal/European best odds to American best odds
    def convert_decimal_to_american(self):
        best_odds = self.best_odds
        for odds in best_odds:
            decimal = odds[ODDS_INDEX]
            if decimal >= 2:
                american = (decimal - 1) * 100
            elif decimal < 2:
                american = -100 / (decimal - 1)
            odds[ODDS_INDEX] = round(american, 2)
        return best_odds
     
    def calculate_arbitrage_bets(self):
        bet_amounts = []
        for outcome in range(self.num_outcomes):
            individual_arbitrage_percentage = 1 / self.best_odds[outcome][ODDS_INDEX]
            bet_amount = (BET_SIZE * individual_arbitrage_percentage) / self.total_arbitrage_percentage
            bet_amounts.append(round(bet_amount, 2))
        
        self.bet_amounts = bet_amounts
        return bet_amounts
    
events = []
for data in odds_response:
    events.append(Event(data))
    #print(data)
    #print()

arbitrage_events = []
for event in events:
    best_odds = event.find_best_odds()
    if event.arbitrage():
        arbitrage_events.append(event)
#print(arbitrage_events)   

for event in arbitrage_events:
    event.calculate_arbitrage_bets()
    event.convert_decimal_to_american()

MAX_OUTCOMES = max([event.num_outcomes for event in arbitrage_events])
ARBITRAGE_EVENTS_COUNT = len(arbitrage_events)


my_columns = ['ID', 'Sport Key', 'Expected Earnings'] + list(np.array([[f'Bookmaker #(outcome)', f'Name #(outcome)', f'Odds #(outcome)', f'Amount to Buy #{(outcome)}'] for outcome in range(1, (MAX_OUTCOMES) + 1)]).flatten())
dataframe = pd.DataFrame(columns=my_columns)

for event in arbitrage_events:
    # print(event.best_odds)
    row = []
    row.append(event.id)
    row.append(event.sport_key)
    row.append(round(event.expected_earnings, 2))
    for index, outcome in enumerate(event.best_odds):
        row.append(outcome[BOOKMAKER_INDEX])
        row.append(outcome[NAME_INDEX])
        row.append(outcome[ODDS_INDEX])
        row.append(event.bet_amounts[index])
    while len(row) < len(dataframe.columns):
        row.append('N/A')
    dataframe.loc[len(dataframe.index)] = row

writer = pd.ExcelWriter('bets.xlsx')
dataframe.to_excel(writer, index=False)
writer.close()

BLACK = '000000'
LIGHT_GREY = 'D6D6D6'
DARK_GREY = '9F9F9F'
RED = 'FEA0A0'
BLUE = 'A0CEFE'
YELLOW = 'FFE540'

COLORS = [RED, BLUE]

ID_COLUMN_FILL = PatternFill(fill_type='solid', start_color=DARK_GREY, end_color=DARK_GREY)
SPORT_KEY_COLUMN_FILL = PatternFill(fill_type='solid', start_color=LIGHT_GREY, end_color=LIGHT_GREY)
EXPECTED_EARNINGS_COLUMN_FILL = PatternFill(fill_type='solid', start_color=YELLOW, end_color=YELLOW)

CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='bottom', indent=0)

TOP_ROW_BORDER = Border(bottom=Side(border_style='thick', color=BLACK))
NORMAL_ROW_BORDER = Border(top=Side(border_style='thin', color=LIGHT_GREY), bottom=Side(border_style='thin', color=DARK_GREY))

wb = load_workbook('bets.xlsx')
ws = wb.active
ws.title = 'Upcoming'
# changing width
for col in range(1, 26):
    ws.column_dimensions[chr(col + 64)].width = 20

for cell in ws['A']:
    cell.fill = ID_COLUMN_FILL
    cell.alignment = CENTER_ALIGNMENT
    
for cell in ws['B']:
    cell.fill = SPORT_KEY_COLUMN_FILL
    cell.alignment = CENTER_ALIGNMENT
    
for cell in ws['C']:
    cell.fill = EXPECTED_EARNINGS_COLUMN_FILL
    cell.alignment = CENTER_ALIGNMENT
    cell.number_format = numbers.BUILTIN_FORMATS[7]

START_INDEX = 'D'
for index in range(MAX_OUTCOMES):
    for col in ws[START_INDEX : chr(ord(START_INDEX) + 3)]:
        for cell in col:
            color = COLORS[int(index % 2)]
            cell.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
            cell.alignment = CENTER_ALIGNMENT
            if cell.column % 4 == 3:
                cell.number_format = numbers.BUILTIN_FORMATS[7]
            
    START_INDEX = chr(ord(START_INDEX) + 4)

for cell in ws['1']:
    cell.border = TOP_ROW_BORDER

for row in range(2, ARBITRAGE_EVENTS_COUNT + 2):
    for cell in ws[str(row)]:
        cell.border = NORMAL_ROW_BORDER
    
wb.save('upcoming_events_bets.xlsx')
